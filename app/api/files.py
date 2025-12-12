# app/api/files.py
import os
import csv
import traceback
from typing import Optional, List

from flask import (
    Blueprint, request, jsonify, current_app, send_file, render_template
)
from app import db
from app.models import UploadedFile
from sqlalchemy import text, or_

# optional Excel preview dependency
try:
    import openpyxl  # type: ignore
    HAVE_OPENPYXL = True
except Exception:
    HAVE_OPENPYXL = False


files_bp = Blueprint("files_api", __name__, url_prefix="/api/files")


def get_uploads_dir() -> str:
    """
    Return absolute path to uploads directory (configurable via UPLOAD_FOLDER).
    Fallback: project_root/uploads
    """
    cfg = current_app.config.get("UPLOAD_FOLDER")
    if cfg:
        return os.path.abspath(cfg)
    project_root = os.path.abspath(os.path.join(current_app.root_path, ".."))
    return os.path.join(project_root, "uploads")


def _normalize_name(name: str) -> str:
    """Return a normalized filename for matching (lower/underscores)."""
    return name.replace(" ", "_").lower()


def find_file_on_disk(uploaded: UploadedFile) -> Optional[str]:
    """
    Heuristic search for the real file on disk for an UploadedFile row.

    Tries:
      - exact file_name or original_file_name
      - "<id>_file_name" and "<id>_original_file_name"
      - normalized variants (spaces->underscores)
      - check "in", endswith, startswith on filenames in uploads dir
    Returns absolute path or None.
    """
    uploads_dir = get_uploads_dir()
    if not os.path.isdir(uploads_dir):
        current_app.logger.debug("Uploads dir does not exist: %s", uploads_dir)
        return None

    candidates: List[str] = []
    # gather possible stored/original names
    if getattr(uploaded, "file_name", None):
        candidates.append(uploaded.file_name)
    if getattr(uploaded, "original_file_name", None):
        candidates.append(uploaded.original_file_name)
    # id prefixed variants
    if getattr(uploaded, "file_name", None):
        candidates.append(f"{uploaded.id}_{uploaded.file_name}")
    if getattr(uploaded, "original_file_name", None):
        candidates.append(f"{uploaded.id}_{uploaded.original_file_name}")

    # normalized variants
    norm_candidates = []
    for c in candidates:
        if not c:
            continue
        norm_candidates.append(c)
        norm_candidates.append(_normalize_name(c))
        # also try with spaces removed
        norm_candidates.append(c.replace(" ", ""))
        norm_candidates.append(_normalize_name(c).replace("_", ""))

    # keep unique
    candidates = []
    seen = set()
    for c in norm_candidates:
        if not c:
            continue
        key = c.lower()
        if key not in seen:
            seen.add(key)
            candidates.append(c)

    # 1) exact path checks
    for cand in candidates:
        p = os.path.join(uploads_dir, cand)
        if os.path.exists(p):
            current_app.logger.debug("find_file_on_disk: found exact path %s", p)
            return p

    # 2) scan directory and try robust matching
    try:
        for fname in os.listdir(uploads_dir):
            fname_lower = fname.lower()
            for cand in candidates:
                cand_lower = cand.lower()
                # exact name match
                if fname_lower == cand_lower:
                    p = os.path.join(uploads_dir, fname)
                    current_app.logger.debug("find_file_on_disk: matched exact filename %s", p)
                    return p
                # endswith match (original file may be prefixed)
                if fname_lower.endswith(cand_lower):
                    p = os.path.join(uploads_dir, fname)
                    current_app.logger.debug("find_file_on_disk: matched endswith %s -> %s", fname, p)
                    return p
                # contains match (less strict)
                if cand_lower in fname_lower:
                    p = os.path.join(uploads_dir, fname)
                    current_app.logger.debug("find_file_on_disk: matched contains %s -> %s", fname, p)
                    return p
                # startswith id (uploaded id prefix)
                if fname_lower.startswith(str(uploaded.id)):
                    p = os.path.join(uploads_dir, fname)
                    current_app.logger.debug("find_file_on_disk: matched startswith id %s -> %s", fname, p)
                    return p
    except Exception:
        current_app.logger.exception("Error listing uploads dir while searching for file")

    current_app.logger.debug("find_file_on_disk: no candidate matched for uploaded id=%s", getattr(uploaded, "id", None))
    return None


@files_bp.route("", methods=["GET"])
def list_files():
    """List uploaded files. query params: q (search), exam_type"""
    try:
        q = (request.args.get("q") or "").strip()
        et = (request.args.get("exam_type") or "").strip()

        query = UploadedFile.query
        if et:
            query = query.filter_by(exam_type=et)
        if q:
            like = f"%{q}%"
            query = query.filter(
                or_(
                    UploadedFile.file_name.ilike(like),
                    UploadedFile.original_file_name.ilike(like)
                )
            )

        items = []
        for f in query.order_by(UploadedFile.uploaded_on.desc()).all():
            items.append({
                "id": f.id,
                "file_name": f.file_name,
                "original_file_name": f.original_file_name,
                "exam_type": f.exam_type,
                "uploaded_on": f.uploaded_on.isoformat() if f.uploaded_on else None,
                "uploaded_by": f.uploaded_by
            })
        return jsonify({"total": len(items), "items": items})
    except Exception as e:
        current_app.logger.exception("list_files error")
        return jsonify({"error": "internal error", "detail": str(e)}), 500


@files_bp.route("/<int:file_id>/preview", methods=["GET"])
def preview_file(file_id):
    """
    Preview file: returns { meta, columns, sample_rows } (best-effort).
    Supports CSV and xlsx/xls (if openpyxl installed).
    """
    f = UploadedFile.query.get(file_id)
    if not f:
        return jsonify({"error": "file not found"}), 404

    path = find_file_on_disk(f)
    if not path:
        return jsonify({"error": "file not found on disk", "checked_dir": get_uploads_dir()}), 404

    ext = os.path.splitext(path)[1].lower()
    meta = {
        "file_name": f.file_name,
        "original_file_name": f.original_file_name,
        "exam_type": f.exam_type,
        "uploaded_on": f.uploaded_on.isoformat() if f.uploaded_on else None,
        "size": os.path.getsize(path)
    }

    try:
        if ext == ".csv":
            with open(path, newline="", encoding="utf-8", errors="replace") as fh:
                sample = fh.read(4096)
                fh.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample) if sample else csv.excel
                except Exception:
                    dialect = csv.excel
                reader = csv.reader(fh, dialect)
                rows = []
                for i, row in enumerate(reader):
                    rows.append(row)
                    if i >= 10:
                        break
                columns = rows[0] if rows else []
                sample_rows = rows[1:6] if len(rows) > 1 else []
            return jsonify({"meta": meta, "columns": columns, "sample_rows": sample_rows})

        elif ext in (".xls", ".xlsx"):
            if not HAVE_OPENPYXL:
                return jsonify({
                    "meta": meta,
                    "error": "Excel preview requires 'openpyxl' package. Install with: pip install openpyxl"
                }), 200
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append([("" if v is None else str(v)) for v in row])
                if i >= 10:
                    break
            columns = rows[0] if rows else []
            sample_rows = rows[1:6] if len(rows) > 1 else []
            return jsonify({"meta": meta, "columns": columns, "sample_rows": sample_rows})

        else:
            return jsonify({"meta": meta, "error": f"Preview not supported for '{ext}'"}), 200

    except Exception as e:
        current_app.logger.exception("preview failed")
        return jsonify({"meta": meta, "error": "preview failed", "detail": str(e)}), 500


@files_bp.route("/<int:file_id>/download", methods=["GET"])
def download_file(file_id):
    f = UploadedFile.query.get(file_id)
    if not f:
        return jsonify({"error": "not found"}), 404

    path = find_file_on_disk(f)
    if not path:
        return jsonify({"error": "file not found on disk", "checked_dir": get_uploads_dir()}), 404

    try:
        # send_file handles file streaming and attachments.
        # download_name used when available (Flask >= 2.0)
        download_name = f.original_file_name or f.file_name or os.path.basename(path)
        return send_file(path, as_attachment=True, download_name=download_name)
    except Exception as e:
        current_app.logger.exception("download failed")
        return jsonify({"error": "download failed", "detail": str(e)}), 500


@files_bp.route("/<int:file_id>", methods=["DELETE"])
def delete_file(file_id):
    """
    Delete an uploaded_files row safely.

    Steps:
      1) Set marks.uploaded_file_id = NULL for any marks referencing this file (prevents FK constraint).
      2) Optionally remove physical file on disk (if ?remove_file=true).
      3) Delete the uploaded_files DB row.
    """
    f = UploadedFile.query.get(file_id)
    if not f:
        return jsonify({"error": "not found"}), 404

    remove_file = str(request.args.get("remove_file", "false")).lower() in ("1", "true", "yes")

    try:
        # 1) Disassociate marks that reference this file
        db.session.execute(
            text("UPDATE marks SET uploaded_file_id = NULL WHERE uploaded_file_id = :fid"),
            {"fid": f.id}
        )

        # 2) Optionally remove the physical file
        if remove_file:
            try:
                file_on_disk = find_file_on_disk(f)
                if file_on_disk and os.path.exists(file_on_disk):
                    os.remove(file_on_disk)
            except Exception as ex:
                # warn but continue - don't block DB deletion if filesystem removal fails
                current_app.logger.warning("Could not remove physical file for uploaded id %s: %s", f.id, ex)

        # 3) Delete DB row
        db.session.delete(f)
        db.session.commit()
        return jsonify({"success": True, "deleted_id": file_id})
    except Exception as e:
        db.session.rollback()
        current_app.logger.exception("delete failed")
        return jsonify({"error": "delete failed", "detail": str(e)}), 500


@files_bp.route("/<int:file_id>/view", methods=["GET"])
def view_file(file_id):
    """Render a full-page table preview for an uploaded file."""
    f = UploadedFile.query.get(file_id)
    if not f:
        return "File not found", 404

    path = find_file_on_disk(f)
    if not path:
        return f"File not found on disk ({get_uploads_dir()})", 404

    ext = os.path.splitext(path)[1].lower()
    rows = []
    try:
        if ext == ".csv":
            with open(path, newline="", encoding="utf-8", errors="replace") as fh:
                reader = csv.reader(fh)
                for i, row in enumerate(reader):
                    rows.append([("" if v is None else str(v)) for v in row])
                    if i > 50:
                        break
        elif ext in (".xls", ".xlsx"):
            if not HAVE_OPENPYXL:
                return "Excel preview requires openpyxl. Install with: pip install openpyxl", 500
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append([("" if v is None else str(v)) for v in row])
                if i > 50:
                    break
        else:
            return f"Preview not supported for {ext}", 400
    except Exception as e:
        current_app.logger.exception("view_file failed")
        return f"Error parsing file: {e}", 500

    if not rows:
        return "No data in file", 200

    columns = rows[0]
    sample_rows = rows[1:]
    return render_template("file_view.html", file=f, columns=columns, rows=sample_rows)


# DEBUG helper (only when app.debug=True) to list disk files to help troubleshooting
@files_bp.route("/_debug/list_disk", methods=["GET"])
def debug_list_disk():
    if not current_app.debug:
        return jsonify({"error": "disabled"}), 403
    uploads_dir = get_uploads_dir()
    out = {"uploads_dir": uploads_dir, "exists": os.path.isdir(uploads_dir), "files": []}
    if os.path.isdir(uploads_dir):
        try:
            out["files"] = sorted(os.listdir(uploads_dir))
        except Exception as e:
            out["err"] = str(e)
    return jsonify(out)
